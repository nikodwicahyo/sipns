"""
Modul: laporan_service.py
Deskripsi: Layanan terstruktur untuk generate laporan nilai (PDF & Excel).

Modul ini menyediakan tiga fungsi utama:

1. ``generate_laporan_pdf(kelas)`` — Rekap nilai per kelas dalam format PDF.
2. ``generate_transkrip_pdf(siswa_id)`` — Transkrip nilai personal siswa.
3. ``export_excel(kelas)`` — Ekspor data nilai ke file ``.xlsx``.

Generator PDF menggunakan **WeasyPrint** (rendering HTML→PDF), sedangkan
ekspor Excel menggunakan **openpyxl** (workbook native). Format styling
diatur via inline CSS (PDF) dan openpyxl styles (Excel).

Catatan teknis:
- ``base_url`` di WeasyPrint WAJIB diset agar asset statis (logo, font)
  dapat di-resolve. Kami pakai ``url_for('static', filename='')`` yang
  me-resolve ke URL static Flask (absolute).
- HTTPException (404, 403) di-re-raise apa adanya — tidak boleh dibungkus
  RuntimeError, karena akan menghilangkan HTTP semantics (lihat juga
  catatan di ``app/blueprints/laporan/routes.py``).
- Model ``Nilai`` dan ``Siswa`` di-import LAZY (di dalam fungsi) untuk
  menghindari circular import (``app.models.nilai`` ←→ ``app.services``).

Author : Niko Dwicahyo
Versi  : 1.0.0
"""
from app.services.nilai_service import hitung_statistik_kelas
from flask import render_template, url_for, has_request_context
from app.utils.time import now_jakarta, current_year_jakarta


def _get_base_url():
    """Resolve base URL absolut untuk WeasyPrint (resolve asset statis).

    Returns:
        str | None: URL absolut ke folder ``static`` jika ada request
        context, atau ``None`` jika dipanggil di luar request (mis.
        dari CLI). WeasyPrint tetap bisa render tanpa base_url, tapi
        gambar/font dari ``static/`` tidak akan ter-load.

    Note:
        Fungsi ini private (prefix ``_``) — tidak untuk konsumsi modul lain.
    """
    try:
        if has_request_context():
            # _external=True → absolute URL (http://host/static/)
            return url_for('static', filename='', _external=True)
    except Exception:
        # url_for bisa raise jika luar app context — fallthrough ke None.
        pass
    return None


def generate_laporan_pdf(kelas=None, template='laporan/rekap_kelas.html',
                        mata_pelajaran=None, guru_id=None, siswa_id=None,
                        status_lulus=None, filter_info=None):
    """Generate PDF laporan rekap nilai siswa.

    Args:
        kelas (str, optional): Nama kelas (mis. ``"X-IPA-1"``) untuk filter
            satu kelas. Jika ``None``, semua kelas akan disertakan dan
            data diurutkan ``(kelas, nama)`` dengan kolom Kelas di PDF.
        template (str, optional): Path template Jinja2 untuk body PDF.
            Default ``'laporan/rekap_kelas.html'``. Bisa di-override untuk
            template alternatif (mis. template bulanan).
        mata_pelajaran (str, optional): Filter berdasarkan mata pelajaran.
            Jika diisi, hanya record ``Nilai`` dengan ``mata_pelajaran``
            yang persis cocok yang akan disertakan. Default ``None``
            (tidak ada filter mapel — semua mapel untuk kelas tersebut).
            Untuk akses guru, route handler akan memaksa parameter ini
            sesuai ``current_user.guru.mata_pelajaran`` agar scope data
            sesuai akun login.
        guru_id (int, optional): Filter berdasarkan guru penginput nilai.
            Default ``None`` (semua guru).
        siswa_id (int, optional): Filter berdasarkan siswa tertentu.
            Default ``None`` (semua siswa di kelas).
        status_lulus (str, optional): Filter kelulusan. Nilai yang valid:
            ``'lulus'`` (status_lulus=True), ``'tidak_lulus'`` (status_lulus=False),
            atau ``None`` (semua status). Default ``None``.
        filter_info (dict, optional): Metadata filter aktif untuk
            ditampilkan di header PDF. Contoh: ``{'guru': 'Budi',
            'siswa': 'Andi', 'status': 'Lulus'}``. Field ``None`` di-skip.

    Returns:
        bytes: Konten file PDF siap-download. Caller (``laporan/routes.py``)
        membungkus dalam ``flask.Response`` dengan ``Content-Disposition``.

    Raises:
        RuntimeError: Jika WeasyPrint gagal render (mis. template syntax
            error, missing asset). HTTPException (404) di-re-raise apa
            adanya — lihat catatan modul.
    """
    import weasyprint
    from werkzeug.exceptions import HTTPException
    # Lazy import untuk menghindari circular import (lihat catatan modul).
    from app import db as _db
    from app.models.nilai import Nilai
    from app.models.siswa import Siswa
    try:
        # Query nilai + siswa dalam satu JOIN dengan eagerload siswa.
        # Eagerload mencegah N+1 query saat template mengakses n.siswa.nama dll.
        query = (
            Nilai.query
            .options(_db.joinedload(Nilai.siswa))
            .join(Siswa, Nilai.siswa_id == Siswa.id)
            .filter(Siswa.deleted_at.is_(None))
        )
        # Filter kelas opsional. None = semua kelas.
        if kelas:
            query = query.filter(Siswa.kelas == kelas)
        # Filter mata pelajaran jika diisi (akses guru → mapel guru saja).
        if mata_pelajaran:
            query = query.filter(Nilai.mata_pelajaran == mata_pelajaran)
        if guru_id is not None:
            query = query.filter(Nilai.guru_id == guru_id)
        if siswa_id is not None:
            query = query.filter(Nilai.siswa_id == siswa_id)
        if status_lulus == 'lulus':
            query = query.filter(Nilai.status_lulus.is_(True))
        elif status_lulus == 'tidak_lulus':
            query = query.filter(Nilai.status_lulus.is_(False))
        # Ordering: bila tanpa kelas, urutkan per kelas lalu nama; kalau
        # ada kelas, urutkan nama saja.
        if kelas:
            data_nilai = query.order_by(Siswa.nama).all()
        else:
            data_nilai = query.order_by(Siswa.kelas, Siswa.nama).all()

        # Statistik agregat (rata-rata, tertinggi, dll) untuk footer PDF.
        statistik = hitung_statistik_kelas(data_nilai)

        # Render template Jinja2 → HTML string. Variabel ``data``,
        # ``statistik``, ``kelas``, ``mata_pelajaran``, ``current_year``,
        # ``tanggal_cetak``, ``filter_info`` tersedia di template.
        html_content = render_template(
            template,
            data=data_nilai,
            statistik=statistik,
            kelas=kelas,
            mata_pelajaran=mata_pelajaran,
            filter_info=filter_info or {},
            current_year=current_year_jakarta(),
            tanggal_cetak=now_jakarta().strftime('%d/%m/%Y %H:%M'),
        )
        # Konversi HTML → PDF. base_url agar static asset (logo) ter-load.
        pdf_bytes = weasyprint.HTML(
            string=html_content,
            base_url=_get_base_url(),
        ).write_pdf()
        return pdf_bytes
    except HTTPException:
        # PENTING: 404 (kelas tidak ditemukan) harus diteruskan apa adanya,
        # JANGAN dibungkus RuntimeError. Lihat debugging_log.md Bug #005.
        raise
    except Exception as e:
        # Error lain (template error, WeasyPrint crash) → RuntimeError
        # dengan pesan informatif. Route handler akan Flash error & redirect.
        raise RuntimeError(f"Gagal generate PDF laporan: {e}") from e


def generate_transkrip_pdf(siswa_id):
    """Generate PDF transkrip nilai personal satu siswa (1 halaman).

    Args:
        siswa_id (int): Primary key siswa. Jika tidak ditemukan, otomatis
            raise 404 (NotFound) yang di-handle Flask errorhandler.

    Returns:
        bytes: Konten file PDF transkrip.

    Raises:
        404 NotFound: Jika ``siswa_id`` tidak ada di DB.
        RuntimeError: Jika render template atau WeasyPrint gagal.
    """
    import weasyprint
    from werkzeug.exceptions import HTTPException
    # Lazy import — lihat generate_laporan_pdf untuk konteks.
    from app.models.nilai import Nilai
    from app.models.siswa import Siswa
    try:
        # get_or_404 otomatis raise NotFound(404) — bukan RuntimeError.
        siswa = Siswa.query.get_or_404(siswa_id)
        data_nilai = Nilai.query.filter_by(siswa_id=siswa_id).all()
        html_content = render_template(
            'laporan/transkrip_siswa.html',
            siswa=siswa,
            data=data_nilai,
            current_year=current_year_jakarta(),
            tanggal_cetak=now_jakarta().strftime('%d/%m/%Y %H:%M'),
        )
        pdf_bytes = weasyprint.HTML(
            string=html_content,
            base_url=_get_base_url(),
        ).write_pdf()
        return pdf_bytes
    except HTTPException:
        # Pertahankan HTTP semantics — JANGAN dibungkus RuntimeError.
        raise
    except Exception as e:
        raise RuntimeError(f"Gagal generate PDF transkrip: {e}") from e


def export_excel(kelas=None, dicetak_oleh=None, mata_pelajaran=None,
                 guru_id=None, siswa_id=None, status_lulus=None,
                 filter_info=None):
    """Ekspor data nilai ke file ``.xlsx`` (Microsoft Excel format).

    Format workbook:
    - Row 1: Judul laporan (merged, bold, font besar).
    - Row 2: Info kelas, mata pelajaran, filter, tanggal, pencetak (merged).
    - Row 3: Baris kosong (spacing).
    - Row 4: Header kolom (bold, fill biru, teks putih).
    - Row 5+: Data nilai (alternating row color abu-abu/putih).
    - Row terakhir: Summary (Rata-rata, Tertinggi, Terendah, % Lulus).

    Args:
        kelas (str, optional): Filter kelas. Jika ``None``, ekspor semua
            kelas yang ada di database.
        dicetak_oleh (str, optional): Nama user yang mencetak (untuk info
            di header). Default ``None`` (tidak ditampilkan).
        mata_pelajaran (str, optional): Filter mata pelajaran. Jika diisi,
            hanya record dengan ``mata_pelajaran`` cocok yang diekspor.
            Untuk akses guru, route handler memaksa parameter ini
            sesuai akun login.
        guru_id (int, optional): Filter berdasarkan guru penginput nilai.
            Default ``None`` (semua guru).
        siswa_id (int, optional): Filter berdasarkan siswa tertentu.
            Default ``None`` (semua siswa).
        status_lulus (str, optional): Filter kelulusan. ``'lulus'`` /
            ``'tidak_lulus'`` / ``None`` (semua).
        filter_info (dict, optional): Metadata filter untuk header Excel.
            Hanya field yang ada & tidak None yang akan ditampilkan.

    Returns:
        bytes: Konten file ``.xlsx`` siap-download. Disimpan ke
        ``BytesIO`` sehingga tidak menulis file ke disk.

    Raises:
        Exception: Error dari openpyxl akan di-propagate apa adanya.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from app import db as _db
    from app.models.nilai import Nilai
    from app.models.siswa import Siswa

    # === 1. Inisialisasi Workbook ===
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Nilai"

    # === 2. Query data ===
    query = (
        Nilai.query
        .options(_db.joinedload(Nilai.siswa))
        .join(Siswa, Nilai.siswa_id == Siswa.id)
        .filter(Siswa.deleted_at.is_(None))
    )
    if kelas:
        query = query.filter(Siswa.kelas == kelas)
    if mata_pelajaran:
        query = query.filter(Nilai.mata_pelajaran == mata_pelajaran)
    if guru_id is not None:
        query = query.filter(Nilai.guru_id == guru_id)
    if siswa_id is not None:
        query = query.filter(Nilai.siswa_id == siswa_id)
    if status_lulus == 'lulus':
        query = query.filter(Nilai.status_lulus.is_(True))
    elif status_lulus == 'tidak_lulus':
        query = query.filter(Nilai.status_lulus.is_(False))
    data_nilai = query.order_by(Siswa.kelas, Siswa.nama).all()

    # === 3. Style constants (dideklarasikan di tengah untuk keterbacaan) ===
    title_font = Font(bold=True, size=14, color="1F4E79")
    info_font = Font(size=10, color="333333")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    summary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    summary_font = Font(bold=True, size=11, color="1F4E79")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # === 4. Definisi kolom header ===
    total_cols = 10
    headers = ["No", "NIS", "Nama", "Kelas", "Mata Pelajaran",
               "Tugas", "UTS", "UAS", "Nilai Akhir", "Status"]

    # === 5. Row 1: Title (merged across all columns) ===
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_text = "LAPORAN REKAP NILAI SISWA"
    if mata_pelajaran:
        title_text += f" - MATA PELAJARAN {mata_pelajaran.upper()}"
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # === 6. Row 2: Info (kelas, mapel, filter tambahan, tanggal, pencetak) ===
    label_kelas = kelas if kelas else "Semua Kelas"
    info_text = f"Kelas: {label_kelas}"
    if mata_pelajaran:
        info_text += f" | Mata Pelajaran: {mata_pelajaran}"
    # Tambahkan info filter tambahan (guru, siswa, status) jika ada.
    if filter_info:
        guru_label = filter_info.get('guru')
        siswa_label = filter_info.get('siswa')
        status_label = filter_info.get('status')
        if guru_label:
            info_text += f" | Guru: {guru_label}"
        if siswa_label:
            info_text += f" | Siswa: {siswa_label}"
        if status_label:
            info_text += f" | Status: {status_label}"
    info_text += f" | Tanggal: {now_jakarta().strftime('%d/%m/%Y %H:%M')}"
    if dicetak_oleh:
        info_text += f" | Dicetak oleh: {dicetak_oleh}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    info_cell = ws.cell(row=2, column=1, value=info_text)
    info_cell.font = info_font
    info_cell.alignment = Alignment(horizontal="left", vertical="center")

    # === 7. Row 3: Blank (spacing visual) ===
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)

    # === 8. Row 4: Column headers ===
    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # === 9. Data rows (from row 5) ===
    data_start_row = header_row + 1
    for i, n in enumerate(data_nilai, 1):
        row = data_start_row + i - 1
        values = [
            i,
            n.siswa.nis if n.siswa else '-',
            n.siswa.nama if n.siswa else '-',
            n.siswa.kelas if n.siswa else '-',
            n.mata_pelajaran,
            float(n.nilai_tugas),
            float(n.nilai_uts),
            float(n.nilai_uas),
            float(n.nilai_akhir) if n.nilai_akhir else '',
            'Lulus' if n.status_lulus else 'Tidak Lulus',
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            # Zebra striping: baris genap abu-abu, ganjil putih.
            if i % 2 == 0:
                cell.fill = alt_fill

    # === 10. Summary row ===
    summary_row = data_start_row + len(data_nilai)
    if data_nilai:
        # Hitung ulang statistik untuk footer summary (jaga konsistensi
        # dengan hitung_statistik_kelas() di nilai_service).
        nilai_akhir_list = [float(n.nilai_akhir) for n in data_nilai if n.nilai_akhir is not None]
        lulus_count = sum(1 for n in data_nilai if n.status_lulus)
        if nilai_akhir_list:
            rata_rata = round(sum(nilai_akhir_list) / len(nilai_akhir_list), 2)
            tertinggi = max(nilai_akhir_list)
            terendah = min(nilai_akhir_list)
            persen_lulus = round((lulus_count / len(nilai_akhir_list)) * 100, 1)
        else:
            rata_rata = tertinggi = terendah = persen_lulus = 0

        # Merge 5 kolom pertama jadi "RINGKASAN" label.
        ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=5)
        summary_label = ws.cell(row=summary_row, column=1, value="RINGKASAN")
        summary_label.font = summary_font
        summary_label.fill = summary_fill
        summary_label.border = thin_border
        # Isi fill di sel-sel yang di-merge.
        for c in range(2, 6):
            ws.cell(row=summary_row, column=c).fill = summary_fill
            ws.cell(row=summary_row, column=c).border = thin_border

        # 4 kolom berikutnya: nilai ringkasan.
        summary_data = [
            ("Rata-rata", rata_rata),
            ("Tertinggi", tertinggi),
            ("Terendah", terendah),
            ("% Lulus", f"{persen_lulus}%"),
        ]
        for j, (_label, val) in enumerate(summary_data):
            col = 6 + j
            cell = ws.cell(row=summary_row, column=col, value=val)
            cell.font = summary_font
            cell.fill = summary_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    # === 11. Auto-fit column widths (default 18, custom untuk kolom panjang) ===
    for col in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['A'].width = 6      # No (pendek)
    ws.column_dimensions['C'].width = 28     # Nama (panjang)
    ws.column_dimensions['E'].width = 22     # Mata Pelajaran (panjang)
    ws.column_dimensions['J'].width = 14     # Status

    # === 12. Freeze panes (scroll-friendly: header tetap terlihat) ===
    ws.freeze_panes = f'A{data_start_row}'

    # === 13. Simpan ke BytesIO dan return bytes ===
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
